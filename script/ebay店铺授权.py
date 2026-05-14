import argparse
import os
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


DEFAULT_URL = "https://auxito.lingxing.com/erp/muser/userManage?wsRefresh=1"
DEFAULT_DRIVER_PATH = r"D:\data\chromedriver.exe"
DEFAULT_WAIT_SECONDS = 15


@dataclass
class AssignTask:
    name: str
    shops: List[str]


class LingxingShopAssigner:
    def __init__(self, driver: webdriver.Chrome, wait_seconds: int = DEFAULT_WAIT_SECONDS, action_delay: float = 1.0):
        self.driver = driver
        self.wait = WebDriverWait(driver, wait_seconds)
        self.action_delay = max(0.2, float(action_delay))

    def run(self, tasks: List[AssignTask]) -> Tuple[List[str], List[str]]:
        success_users: List[str] = []
        failed_users: List[str] = []
        for task in tasks:
            print(f"\n====== Processing user: {task.name}, shops: {len(task.shops)} ======")
            last_exc: Optional[Exception] = None
            for attempt in range(1, 3):
                try:
                    self.driver.get(DEFAULT_URL)
                    wait_for_user_manage_ready(self.driver, timeout_seconds=20)
                    self.assign_user_shops(task)
                    last_exc = None
                    break
                except Exception as exc:  # noqa: BLE001
                    last_exc = exc
                    if attempt == 1 and self._is_retryable_error(str(exc)):
                        print(f"[WARN] transient failure, retry once: {exc}")
                        self._pause(1.2)
                        continue
                    break
            if last_exc is not None:
                print(f"[FAILED] {task.name}: {last_exc}")
                failed_users.append(task.name)
                continue
            print(f"[OK] {task.name}")
            success_users.append(task.name)
        return success_users, failed_users

    def assign_user_shops(self, task: AssignTask) -> None:
        self.search_user(task.name)
        row = self.ensure_single_exact_row(task.name)
        self.select_row(row)
        self.open_assign_shop_dialog()
        self.choose_ebay_tab()
        self.select_authorized_shops(task.shops)
        self.confirm_dialog()
        self.wait_dialog_closed()

    def search_user(self, name: str) -> None:
        # Compatible with different Lingxing page variants.
        input_selectors = [
            (By.XPATH, "//input[contains(@placeholder,'姓名')]"),
            (By.XPATH, "//input[contains(@placeholder,'用户')]"),
            (By.XPATH, "//input[contains(@placeholder,'成员')]"),
            (By.XPATH, "//input[contains(@placeholder,'昵称')]"),
            (By.XPATH, "//input[contains(@placeholder,'手机')]"),
            (By.XPATH, "//input[contains(@placeholder,'邮箱')]"),
            (By.XPATH, "//input[contains(@placeholder,'关键字')]"),
            (By.XPATH, "//input[contains(@placeholder,'搜索')]"),
            (By.XPATH, "//input[contains(@class,'el-input__inner') and @inelement='0']"),
            (By.XPATH, "//input[@type='text' and not(@disabled)]"),
        ]
        search_inputs = self._find_all_visible(input_selectors)
        if not search_inputs:
            raise RuntimeError("Search input not found (name/user/member).")

        last_error = None
        for search_input in search_inputs:
            if not self._is_editable_input(search_input):
                continue
            try:
                search_input.click()
                try:
                    search_input.clear()
                except Exception:  # noqa: BLE001
                    pass
                search_input.send_keys(Keys.CONTROL, "a")
                search_input.send_keys(Keys.BACKSPACE)
                search_input.send_keys(name)
                # Prefer explicit search button click; fallback to Enter key.
                clicked = self._try_click_first([
                    (By.XPATH, "//button[normalize-space()='搜索']"),
                    (By.XPATH, "//span[normalize-space()='搜索']/ancestor::button[1]"),
                    (By.XPATH, "//button[normalize-space()='查询']"),
                    (By.XPATH, "//span[normalize-space()='查询']/ancestor::button[1]"),
                    (By.XPATH, "//button[contains(@class,'search')]"),
                ])
                if not clicked:
                    search_input.send_keys(Keys.ENTER)
                time.sleep(0.8)
                return
            except Exception as exc:  # noqa: BLE001
                last_error = exc
                continue

        raise RuntimeError(f"Failed to enter search text in any editable input. Last error: {last_error}")

    def ensure_single_exact_row(self, target_name: str):
        rows = self._get_table_rows()
        if len(rows) != 1:
            raise RuntimeError(f"Search result count is not 1, actual: {len(rows)}")

        name_cell = self._find_name_cell(rows[0])
        row_name = name_cell.text.strip() if name_cell is not None else rows[0].text.strip().split("\n")[0]
        row_text = rows[0].text.strip()
        if row_name != target_name and target_name not in row_text:
            raise RuntimeError(f"Matched user name mismatch, expected: {target_name}, actual: {row_name}")
        return rows[0]

    def select_row(self, row) -> None:
        checkbox_candidates = [
            ".//label[contains(@class,'el-checkbox') or contains(@class,'checkbox')]",
            ".//*[contains(@class,'el-checkbox__inner')]",
            ".//input[@type='checkbox']/ancestor::*[self::label or self::span or self::div][1]",
            ".//td[1]//*[contains(@class,'checkbox')]",
            ".//td[1]",
        ]
        for xpath in checkbox_candidates:
            try:
                elements = row.find_elements(By.XPATH, xpath)
            except Exception:  # noqa: BLE001
                elements = []
            for element in elements:
                if not element.is_displayed():
                    continue
                try:
                    self._safe_click(element)
                    time.sleep(0.3)
                    return
                except Exception:  # noqa: BLE001
                    continue
        # Last fallback: click the row itself (some tables toggle selection by row click).
        try:
            self._safe_click(row)
            time.sleep(0.3)
            return
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError("Row checkbox not found.") from exc

    def open_assign_shop_dialog(self) -> None:
        self._click_first([
            # User-confirmed DOM: <span class="el-tooltip text-wrap"><span>批量</span>...</span>
            (By.XPATH, "//span[contains(@class,'el-tooltip') and contains(@class,'text-wrap')][.//span[normalize-space()='批量']]"),
            (By.XPATH, "//span[normalize-space()='批量']/ancestor::span[contains(@class,'el-tooltip')][1]"),
            (By.XPATH, "//span[normalize-space()='批量']/ancestor::*[contains(@class,'ak-drop-button') or contains(@class,'el-dropdown')][1]"),
            (By.XPATH, "//button[.//span[normalize-space()='批量'] or normalize-space()='批量']"),
            (By.XPATH, "//span[normalize-space()='批量']/ancestor::button[1]"),
        ], "批量按钮")
        self._pause(1.0)
        self._click_first([
            # User-confirmed DOM: <div class="el-tooltip"><span>分配店铺</span></div>
            (By.XPATH, "//div[contains(@class,'el-tooltip')][.//span[normalize-space()='分配店铺']]"),
            (By.XPATH, "//span[normalize-space()='分配店铺']/ancestor::div[contains(@class,'el-tooltip')][1]"),
            (By.XPATH, "//*[self::li or self::a or self::span][normalize-space()='分配店铺']"),
        ], "分配店铺菜单")
        self.wait.until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(@class,'dialog') or @role='dialog']"))
        )
        self.wait.until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(@class,'platform-item')]"))
        )
        self._pause(1.2)

    def choose_ebay_tab(self) -> None:
        # User-confirmed DOM:
        # <div class="platform-item"><span ...>eBay</span>...</div>
        ebay_xpath = (
            "//div[contains(@class,'platform-item')]"
            "[.//span[contains(translate(normalize-space(.),'EBAY','ebay'),'ebay')]]"
        )
        try:
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, ebay_xpath))
            )
        except TimeoutException:
            pass
        ebay_items = self.driver.find_elements(By.XPATH, ebay_xpath)
        if not ebay_items:
            raise RuntimeError("eBay platform item not found.")

        clicked = False
        # Click the last visible element first; it is usually the top-most popup layer.
        visible_items = [item for item in ebay_items if item.is_displayed()]
        for item in reversed(visible_items):
            if not item.is_displayed():
                continue
            try:
                self._safe_click(item)
                clicked = True
            except Exception:  # noqa: BLE001
                try:
                    span = item.find_element(By.XPATH, ".//span[normalize-space()='eBay']")
                    self._safe_click(span)
                    clicked = True
                except Exception:  # noqa: BLE001
                    self.driver.execute_script("arguments[0].click();", item)
                    clicked = True
            if clicked:
                break
        if not clicked:
            raise RuntimeError("Failed to click eBay platform item.")

        self._pause(1.2)
        dialog = None
        try:
            dialog = self._find_assign_dialog()
        except Exception:  # noqa: BLE001
            dialog = None
        if dialog is not None and not self._is_ebay_tab_active(dialog):
            # A second click in the same dialog usually fixes missed click state.
            for item in reversed(visible_items):
                if not item.is_displayed():
                    continue
                self.driver.execute_script("arguments[0].click();", item)
                self._pause(0.8)
                if self._is_ebay_tab_active(dialog):
                    break
        # Avoid operating too early: wait until shop list appears and stabilizes.
        self._wait_shop_list_ready(timeout_seconds=15)

    def select_authorized_shops(self, shops: List[str]) -> None:
        for shop in shops:
            normalized = shop.strip()
            if not normalized:
                continue
            self._search_shop_in_dialog(normalized)
            group = self._find_shop_group()
            clicked = self._click_label_by_text(group, normalized)
            if not clicked:
                visible_titles = self._get_visible_shop_titles(group, max_items=20)
                raise RuntimeError(
                    f"Shop checkbox not found: {normalized}. "
                    f"Sample visible shops: {visible_titles}"
                )
            self._pause(0.5)

    def confirm_dialog(self) -> None:
        self._click_first([
            # User-confirmed DOM:
            # <button class="el-button el-button--primary el-button--small is-round">...<span class="el-tooltip text-wrap">确定</span></button>
            (By.XPATH, "//button[contains(@class,'el-button--primary') and contains(@class,'el-button--small') and contains(@class,'is-round')][.//span[contains(@class,'el-tooltip') and contains(@class,'text-wrap') and normalize-space()='确定']]"),
            (By.XPATH, "//span[contains(@class,'el-tooltip') and contains(@class,'text-wrap') and normalize-space()='确定']/ancestor::button[1]"),
            (By.XPATH, "//button[normalize-space()='确定']"),
            (By.XPATH, "//span[normalize-space()='确定']/ancestor::button[1]"),
        ], "确定按钮")

    def wait_dialog_closed(self) -> None:
        try:
            self.wait.until_not(
                EC.presence_of_element_located(
                    (By.XPATH, "//div[contains(@class,'dialog') and not(contains(@style,'display: none'))]")
                )
            )
        except TimeoutException:
            # 部分页面关闭慢，容忍并继续
            time.sleep(1.0)

    def _get_table_rows(self):
        row_selectors = [
            (By.XPATH, "//div[contains(@class,'el-table__body-wrapper')]//tbody/tr[not(contains(@style,'display: none'))]"),
            (By.XPATH, "//tr[contains(@class,'el-table__row') and not(contains(@style,'display: none'))]"),
            (By.XPATH, "//table//tbody/tr[not(contains(@style,'display: none'))]"),
            (By.XPATH, "//tr[.//td and not(ancestor::thead)]"),
        ]
        for by, locator in row_selectors:
            rows = [r for r in self.driver.find_elements(by, locator) if r.is_displayed()]
            if rows:
                return rows
        raise RuntimeError("Data table rows not found.")

    @staticmethod
    def _find_name_cell(row):
        candidates = row.find_elements(By.XPATH, ".//td")
        if not candidates:
            return None
        # 优先返回有文本的第一个单元格（通常是姓名列）
        for cell in candidates:
            text = cell.text.strip()
            if text:
                return cell
        return candidates[0]

    def _find_assign_dialog(self):
        selectors = [
            (
                By.XPATH,
                "//div[(contains(@class,'dialog') or @role='dialog') and not(contains(@style,'display: none'))]",
            ),
            (
                By.XPATH,
                "//div[contains(@class,'el-dialog__wrapper') and not(contains(@style,'display: none'))]",
            ),
        ]
        dialog = self._find_first_visible(selectors)
        if dialog is None:
            raise RuntimeError("Assign-shop dialog not found.")
        return dialog

    def _find_shop_group(self):
        dialog = self._find_assign_dialog()
        selectors = [
            (By.XPATH, ".//div[contains(@class,'el-checkbox-group') and contains(@class,'seller-wrapper')]"),
            (By.XPATH, ".//div[@role='group' and contains(@class,'el-checkbox-group')]"),
        ]
        group = self._find_first_visible_in_root(dialog, selectors)
        if group is None:
            raise RuntimeError("Shop checkbox group not found.")
        return group

    def _is_ebay_tab_active(self, root=None) -> bool:
        if root is None:
            root = self.driver
        # We don't fully know the active class name in this page version,
        # so check common active-state class tokens around the eBay platform-item.
        active_tokens = ("active", "selected", "is-active", "is-selected", "current")
        candidates = root.find_elements(
            By.XPATH,
            ".//div[contains(@class,'platform-item')][.//span[contains(translate(normalize-space(.),'EBAY','ebay'),'ebay')]]",
        )
        for item in candidates:
            if not item.is_displayed():
                continue
            class_name = (item.get_attribute("class") or "").lower()
            if any(token in class_name for token in active_tokens):
                return True
        return False

    def _search_shop_in_dialog(self, shop_name: str) -> None:
        # User-confirmed DOM: <div class="search-input ..."><input class="el-input__inner" inelement="0"></div>
        dialog = self._find_assign_dialog()
        selectors = [
            (
                By.XPATH,
                ".//div[contains(@class,'search-input') and .//input[contains(@class,'el-input__inner') and @inelement='0']]"
                "//input[contains(@class,'el-input__inner') and @inelement='0' and not(@readonly)]",
            ),
            (
                By.XPATH,
                ".//input[contains(@class,'el-input__inner') and @inelement='0' and not(@readonly)]",
            ),
        ]
        search_input = self._find_first_visible_in_root(dialog, selectors)
        if search_input is None:
            return
        try:
            search_input.click()
            try:
                search_input.clear()
            except Exception:  # noqa: BLE001
                pass
            search_input.send_keys(Keys.CONTROL, "a")
            search_input.send_keys(Keys.BACKSPACE)
            search_input.send_keys(shop_name)
            search_input.send_keys(Keys.ENTER)
            self._pause(0.8)
        except Exception:  # noqa: BLE001
            # Fallback to non-filter path.
            return

    def _click_label_by_text(self, group, text: str) -> bool:
        target = text.strip().upper()
        if not target:
            return False

        labels = group.find_elements(By.XPATH, ".//label[contains(@class,'el-checkbox')]")

        def _match_current_dom() -> bool:
            for label in labels:
                title = (label.get_attribute("title") or "").strip().upper()
                label_text = (
                    (label.text or "")
                    .replace("\n", " ")
                    .replace("\xa0", " ")
                    .strip()
                    .upper()
                )
                if title == target or target in label_text:
                    classes = (label.get_attribute("class") or "").strip()
                    if "is-checked" in classes:
                        return True
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", label)
                    self._safe_click(label)
                    return True
            return False

        # Try current viewport first.
        if _match_current_dom():
            return True

        # If not found, scroll seller list (virtual list scenario) and retry.
        try:
            scroll_height = int(self.driver.execute_script("return arguments[0].scrollHeight || 0;", group))
            client_height = int(self.driver.execute_script("return arguments[0].clientHeight || 0;", group))
        except Exception:  # noqa: BLE001
            scroll_height, client_height = 0, 0

        if scroll_height > client_height and client_height > 0:
            step = max(120, int(client_height * 0.8))
            scroll_top = 0
            max_rounds = 80
            for _ in range(max_rounds):
                labels = group.find_elements(By.XPATH, ".//label[contains(@class,'el-checkbox')]")
                self.driver.execute_script("arguments[0].scrollTop = arguments[1];", group, scroll_top)
                self._pause(0.25)
                if _match_current_dom():
                    return True
                scroll_top += step
                if scroll_top > scroll_height:
                    break
        return False

    @staticmethod
    def _is_retryable_error(msg: str) -> bool:
        retry_tokens = (
            "Data table rows not found",
            "eBay platform item not found",
            "Assign-shop dialog not found",
            "Search result count is not 1",
        )
        return any(token in msg for token in retry_tokens)

    def _wait_shop_list_ready(self, timeout_seconds: int = 15) -> None:
        start_time = time.time()
        stable_rounds = 0
        last_count = -1
        while time.time() - start_time < timeout_seconds:
            try:
                group = self._find_shop_group()
                count = len(group.find_elements(By.XPATH, ".//label[contains(@class,'el-checkbox')]"))
            except Exception:  # noqa: BLE001
                count = 0
            if count > 0 and count == last_count:
                stable_rounds += 1
            else:
                stable_rounds = 0
            if count > 0 and stable_rounds >= 2:
                self._pause(0.4)
                return
            last_count = count
            self._pause(0.4)

    def _pause(self, seconds: float) -> None:
        time.sleep(self.action_delay * seconds)

    @staticmethod
    def _get_visible_shop_titles(group, max_items: int = 20) -> List[str]:
        titles: List[str] = []
        labels = group.find_elements(By.XPATH, ".//label[contains(@class,'el-checkbox')]")
        for label in labels:
            if not label.is_displayed():
                continue
            title = (label.get_attribute("title") or "").strip()
            if not title:
                title = (label.text or "").replace("\n", " ").strip()
            if not title:
                continue
            titles.append(title)
            if len(titles) >= max_items:
                break
        return titles

    def _click_first(self, selectors: List[Tuple[str, str]], alias: str) -> None:
        element = self._find_first_visible(selectors)
        if element is None:
            raise RuntimeError(f"未找到{alias}")
        self._safe_click(element)

    def _try_click_first(self, selectors: List[Tuple[str, str]]) -> bool:
        element = self._find_first_visible(selectors)
        if element is None:
            return False
        self._safe_click(element)
        return True

    def _find_first_visible(self, selectors: List[Tuple[str, str]]):
        for by, locator in selectors:
            elements = self.driver.find_elements(by, locator)
            for ele in elements:
                if ele.is_displayed():
                    return ele
        return None

    @staticmethod
    def _find_first_visible_in_root(root, selectors: List[Tuple[str, str]]):
        for by, locator in selectors:
            elements = root.find_elements(by, locator)
            for ele in elements:
                if ele.is_displayed():
                    return ele
        return None

    def _find_all_visible(self, selectors: List[Tuple[str, str]]):
        visible = []
        seen = set()
        for by, locator in selectors:
            elements = self.driver.find_elements(by, locator)
            for ele in elements:
                if not ele.is_displayed():
                    continue
                element_id = ele.id
                if element_id in seen:
                    continue
                seen.add(element_id)
                visible.append(ele)
        return visible

    @staticmethod
    def _is_editable_input(element) -> bool:
        if not element.is_enabled():
            return False
        readonly = (element.get_attribute("readonly") or "").strip().lower()
        if readonly in {"readonly", "true"}:
            return False
        input_type = (element.get_attribute("type") or "").strip().lower()
        if input_type in {"hidden", "file"}:
            return False
        return True

    def _safe_click(self, element) -> None:
        try:
            self.wait.until(lambda d: element.is_displayed() and element.is_enabled())
            ActionChains(self.driver).move_to_element(element).pause(0.1).click(element).perform()
            return
        except Exception:  # noqa: BLE001
            pass
        try:
            element.click()
            return
        except Exception:  # noqa: BLE001
            pass
        self.driver.execute_script("arguments[0].click();", element)


def _xpath_text_literal(text: str) -> str:
    if "'" not in text:
        return f"'{text}'"
    if '"' not in text:
        return f'"{text}"'
    parts = text.split("'")
    return "concat(" + ", \"'\", ".join(f"'{part}'" for part in parts) + ")"


def resolve_default_excel() -> Path:
    user_home = Path.home()
    candidates = [
        Path(r"D:\data\ebay店铺授权.xlsx"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def parse_excel_tasks(excel_path: Path) -> List[AssignTask]:
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    if not rows:
        raise RuntimeError("Excel header row is empty.")

    raw_header = [str(col).strip() if col is not None else "" for col in rows[0]]
    normalized_header = [_normalize_header_text(col) for col in raw_header]
    name_idx = _find_header_index(normalized_header, {"姓名", "名字", "用户姓名", "真实姓名"}, ["姓名"])
    shops_idx = _find_header_index(normalized_header, {"授权店铺", "店铺", "授权店铺列表"}, ["店铺", "授权"])
    if name_idx is None or shops_idx is None:
        # 如果表头不规范，则回退到前两列，兼容“无表头/别名表头”场景
        print(f"[INFO] Standard headers not detected, fallback to column1=name, column2=shops. Raw headers: {raw_header}")
        name_idx, shops_idx = 0, 1

    tasks: List[AssignTask] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name_val = "" if name_idx >= len(row) or row[name_idx] is None else str(row[name_idx]).strip()
        shops_val = "" if shops_idx >= len(row) or row[shops_idx] is None else str(row[shops_idx]).strip()
        if not name_val:
            continue
        shops = _split_shops(shops_val)
        if not shops:
            print(f"[SKIP] {name_val} has no authorized shops.")
            continue
        tasks.append(AssignTask(name=name_val, shops=shops))

    wb.close()
    if not tasks:
        raise RuntimeError("No valid rows parsed from Excel. Please check the sheet content.")
    return tasks


def _find_header_index(headers: List[str], accepted: set, fuzzy_keywords: Optional[List[str]] = None) -> Optional[int]:
    for idx, col in enumerate(headers):
        if col in accepted:
            return idx
    if fuzzy_keywords:
        for idx, col in enumerate(headers):
            if any(keyword in col for keyword in fuzzy_keywords):
                return idx
    return None


def _normalize_header_text(text: str) -> str:
    # 仅保留中英文与数字，增强对 "*真实姓名" 这类字段的识别
    return re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "", text or "")


def _split_shops(raw_text: str) -> List[str]:
    if not raw_text:
        return []
    parts = re.split(r"[,，]", raw_text)
    cleaned = [p.strip() for p in parts if p and p.strip()]
    # 去重且保留顺序
    unique: List[str] = []
    seen = set()
    for item in cleaned:
        if item in seen:
            continue
        seen.add(item)
        unique.append(item)
    return unique


def build_driver(chromedriver_path: str, user_data_dir: Optional[str] = None) -> webdriver.Chrome:
    if not os.path.exists(chromedriver_path):
        raise FileNotFoundError(f"chromedriver not found: {chromedriver_path}")

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--ignore-certificate-errors")
    if user_data_dir:
        Path(user_data_dir).mkdir(parents=True, exist_ok=True)
        options.add_argument(f"--user-data-dir={user_data_dir}")

    service = Service(executable_path=chromedriver_path)
    return webdriver.Chrome(service=service, options=options)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Batch assign eBay shops to Lingxing ERP users.")
    parser.add_argument("--excel", default=str(resolve_default_excel()), help="Excel path. Default tries Documents/a.xlsx")
    parser.add_argument("--driver", default=DEFAULT_DRIVER_PATH, help="Path to chromedriver.exe")
    parser.add_argument("--url", default=DEFAULT_URL, help="Lingxing user-management URL")
    parser.add_argument("--user-data-dir", default=str(Path("D:/data/chromedata/lingxing_manual_login")), help="Chrome user data directory")
    parser.add_argument("--page-timeout", type=int, default=30, help="Max seconds to wait for page ready")
    parser.add_argument("--action-delay", type=float, default=1.2, help="Global action delay multiplier (higher = slower)")
    return parser.parse_args()


def wait_for_user_manage_ready(driver: webdriver.Chrome, timeout_seconds: int) -> None:
    start_ts = time.time()
    while True:
        if is_user_manage_page_ready(driver):
            return
        elapsed = int(time.time() - start_ts)
        if elapsed >= timeout_seconds:
            raise TimeoutError(
                f"Page ready timeout ({timeout_seconds}s). "
                "Please confirm the current page is user management."
            )
        time.sleep(1.0)


def is_user_manage_page_ready(driver: webdriver.Chrome) -> bool:
    try:
        current_url = driver.current_url or ""
        if "muser/userManage" not in current_url:
            return False
        ready_markers = [
            (By.XPATH, "//button[.//span[normalize-space()='批量'] or normalize-space()='批量']"),
            (By.XPATH, "//table[contains(@class,'el-table')]"),
            (By.XPATH, "//tbody/tr"),
        ]
        for by, locator in ready_markers:
            elements = driver.find_elements(by, locator)
            if any(ele.is_displayed() for ele in elements):
                return True
        return False
    except Exception:  # noqa: BLE001
        return False


def main() -> None:
    args = parse_args()
    excel_path = Path(args.excel)
    tasks = parse_excel_tasks(excel_path)
    print(f"Excel parsed: {len(tasks)} users to process.")

    driver = build_driver(args.driver, args.user_data_dir)
    try:
        driver.get(args.url)
        wait_for_user_manage_ready(driver, args.page_timeout)

        assigner = LingxingShopAssigner(driver, action_delay=args.action_delay)
        success_users, failed_users = assigner.run(tasks)
        print("\n====== Finished ======")
        print(f"Success: {len(success_users)}")
        print(f"Failed: {len(failed_users)}")
        if failed_users:
            print("Failed users:")
            for name in failed_users:
                print(f"- {name}")
    finally:
        print("\nScript done. Browser will close in 10 seconds...")
        time.sleep(10)
        driver.quit()


if __name__ == "__main__":
    main()
