import os
import re
import time
import traceback
import sys
sys.path.append(r"D:\app\astpy")
import openpyxl
from selenium.common import StaleElementReferenceException, TimeoutException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

from common.ResultObj import ResultObj
from rpa.RpaAdmin import RpaAdmin
from utils.BrowserUtils import BrowserUtils


class DeleteLxEbayPairedMskuJob(RpaAdmin):
    EXCEL_FILE = r"D:\data\eBay删除已配对msku店铺清单.xlsx"
    PAIR_LIST_URL = "https://auxito.lingxing.com/erp/mmulti/pairList"
    PAGE_SIZE = 200
    MAX_BROWSER_RECOVERY = 3

    def __init__(self, jobname):
        super().__init__(jobname)
        self.jobname = jobname

    def action(self, **kwargs):
        browser = None
        try:
            store_list = self.read_store_list(kwargs.get("excel_file", self.EXCEL_FILE))
            if len(store_list) == 0:
                return ResultObj.error(ResultObj.INVALID_INPUT, "Excel中没有可处理的店铺简称")

            browser = self.open_pair_list_page()

            total_deleted = 0
            for idx, store_short_name in enumerate(store_list, start=1):
                store_name = f"[eBay].{store_short_name}"
                self.logger.info(f"开始处理第{idx}/{len(store_list)}个店铺：{store_name}")
                store_deleted = 0
                recovery_times = 0
                while True:
                    try:
                        deleted_count = self.delete_store_pair_data(browser, store_name)
                        store_deleted += deleted_count
                        break
                    except Exception as exc:
                        if not self.is_browser_session_dead(exc):
                            raise
                        recovery_times += 1
                        if recovery_times > self.MAX_BROWSER_RECOVERY:
                            raise
                        self.logger.warning(
                            f"浏览器会话中断，重新打开后继续处理店铺：{store_name}，"
                            f"第{recovery_times}/{self.MAX_BROWSER_RECOVERY}次"
                        )
                        try:
                            browser.quit()
                        except Exception:
                            pass
                        browser = self.open_pair_list_page()
                deleted_count = store_deleted
                total_deleted += deleted_count
                self.logger.info(f"店铺：{store_name} 删除完成，本店铺删除{deleted_count}条")

            self.logger.info(f"所有店铺处理完成，共删除{total_deleted}条，等待120秒后退出程序")
            time.sleep(120)
            return ResultObj.success(msg=f"处理完成，共删除{total_deleted}条")
        except Exception:
            return ResultObj.error(ResultObj.FATAL_ERROR, traceback.format_exc())
        finally:
            if browser is not None:
                try:
                    browser.quit()
                except Exception:
                    pass

    def open_pair_list_page(self):
        code, msg, browser = BrowserUtils.openChrome(
            chromdatapath=self.admin.getChromeUserPath(f"{self.__class__.__name__}.action")
        )
        if code != 0:
            raise RuntimeError(msg)

        browser.get(self.PAIR_LIST_URL)
        if not self.wait_login_and_page_ready(browser):
            try:
                browser.quit()
            except Exception:
                pass
            raise TimeoutException("120秒内没有检测到领星已配对页面，请确认是否已完成登录")

        self.close_popups(browser)
        self.ensure_paired_tab(browser)
        self.ensure_page_size_200(browser)
        return browser

    def is_browser_session_dead(self, exc):
        msg = (str(exc) or "").lower()
        needles = (
            "invalid session id",
            "session deleted",
            "disconnected",
            "not connected to devtools",
            "chrome not reachable",
            "target window already closed",
            "browser has closed",
            "no such session",
        )
        return any(needle in msg for needle in needles)

    def read_store_list(self, excel_file):
        if not os.path.exists(excel_file):
            raise FileNotFoundError(f"Excel文件不存在：{excel_file}")

        workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            store_list = []
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
                value = row[0].value
                if value is None:
                    continue
                store_short_name = str(value).strip()
                if store_short_name == "" or store_short_name == "店铺简称":
                    continue
                store_list.append(store_short_name)
            return store_list
        finally:
            workbook.close()

    def wait_login_and_page_ready(self, browser):
        end_time = time.time() + 120
        tab_xpath = "//div[@id='tab-1' and @role='tab' and normalize-space()='已配对']"
        while time.time() < end_time:
            try:
                if len(browser.find_elements(By.XPATH, tab_xpath)) > 0:
                    self.wait_page_idle(browser)
                    return True
            except WebDriverException:
                pass
            time.sleep(1)
        return False

    def close_popups(self, browser):
        close_xpaths = [
            "//span[contains(text(),'平台公告')]/../..//button[normalize-space()='关闭']",
            "//button[normalize-space()='关闭']",
            "//button[normalize-space()='确 认']",
            "//button[normalize-space()='确认']",
        ]
        for close_xpath in close_xpaths:
            try:
                for elem in browser.find_elements(By.XPATH, close_xpath):
                    if elem.is_displayed() and elem.is_enabled():
                        browser.execute_script("arguments[0].click();", elem)
                        time.sleep(0.5)
            except WebDriverException:
                pass

    def ensure_paired_tab(self, browser):
        tab = WebDriverWait(browser, 20).until(
            EC.presence_of_element_located((By.XPATH, "//div[@role='tab' and normalize-space()='已配对']"))
        )
        if "is-active" not in (tab.get_attribute("class") or ""):
            browser.execute_script("arguments[0].click();", tab)
            self.wait_page_idle(browser)
            time.sleep(1)

    def ensure_page_size_200(self, browser):
        if self.get_current_page_size(browser) == self.PAGE_SIZE:
            return

        try:
            page_size_select = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located((
                    By.XPATH,
                    "(//*[contains(normalize-space(),'条/页')]/ancestor::div[contains(@class,'el-select')][1])[last()]"
                ))
            )
            browser.execute_script("arguments[0].scrollIntoView({block:'center'}); arguments[0].click();", page_size_select)
            option = WebDriverWait(browser, 10).until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    f"(//li[contains(@class,'el-select-dropdown__item') and contains(normalize-space(),'{self.PAGE_SIZE}条/页')])[last()]"
                ))
            )
            browser.execute_script("arguments[0].click();", option)
            self.wait_page_idle(browser)
            time.sleep(1)
        except TimeoutException:
            self.logger.warning("未能调整为200条/页，继续使用当前页面行数执行")

    def get_current_page_size(self, browser):
        try:
            elems = browser.find_elements(By.XPATH, "//span[contains(normalize-space(),'条/页')]")
            for elem in elems:
                text = (elem.text or elem.get_attribute("textContent") or "").strip()
                if f"{self.PAGE_SIZE}条/页" in text:
                    return self.PAGE_SIZE
        except WebDriverException:
            pass
        return None

    def delete_store_pair_data(self, browser, store_name):
        self.search_store(browser, store_name)
        deleted_count = 0
        empty_retry = 0

        while True:
            self.wait_page_idle(browser)
            current_count = self.get_result_count(browser)
            if current_count <= 0:
                empty_retry += 1
                if empty_retry >= 2:
                    return deleted_count
                time.sleep(1)
                continue

            empty_retry = 0
            batch_count = min(current_count, self.PAGE_SIZE)
            self.logger.info(f"店铺：{store_name} 当前查询到{current_count}条，准备删除本页{batch_count}条")
            self.select_all_current_page(browser)
            self.click_delete_button(browser)
            self.confirm_delete(browser)
            self.wait_deleted(browser, current_count)
            deleted_count += batch_count

        pass

    def search_store(self, browser, store_name):
        last_exc = None
        for attempt in range(1, 4):
            try:
                select_root = self.find_store_select_root(browser)
                self.select_store_from_dropdown(browser, select_root, store_name)
                break
            except TimeoutException as exc:
                last_exc = exc
                self.logger.warning(f"店铺筛选第{attempt}次失败，准备重新打开店铺下拉：{exc}")
                self.close_dropdown(browser)
                time.sleep(1)
        else:
            raise last_exc
        self.click_query_button_if_exists(browser)
        self.wait_page_idle(browser)
        time.sleep(1)

    def select_store_from_dropdown(self, browser, select_root, store_name):
        search_input = None
        for _ in range(3):
            self.open_store_dropdown(browser, select_root)
            try:
                search_input = self.find_dropdown_search_input(browser)
                break
            except TimeoutException:
                time.sleep(0.5)
        if search_input is None:
            raise TimeoutException("无法定位店铺下拉搜索输入框")
        self.uncheck_selected_store_options(browser)
        self.input_dropdown_search_text(browser, search_input, store_name)
        option = WebDriverWait(browser, 10).until(
            EC.element_to_be_clickable((
                By.XPATH,
                f"//li[contains(@class,'el-select-dropdown__item') and .//*[normalize-space()='{store_name}' or contains(normalize-space(),'{store_name}')]]"
                f"|//li[contains(@class,'el-select-dropdown__item') and (normalize-space()='{store_name}' or contains(normalize-space(),'{store_name}'))]"
            ))
        )
        browser.execute_script("arguments[0].click();", option)
        time.sleep(0.5)
        self.confirm_dropdown_filter(browser)

    def uncheck_selected_store_options(self, browser):
        unchecked_count = 0
        for _ in range(20):
            checked_item = self.find_checked_dropdown_option(browser)
            if checked_item is None:
                break
            browser.execute_script("arguments[0].click();", checked_item)
            unchecked_count += 1
            time.sleep(0.2)
        if unchecked_count > 0:
            self.logger.info(f"已取消店铺下拉中旧的勾选项：{unchecked_count}个")

    def find_checked_dropdown_option(self, browser):
        option_xpaths = [
            "(//div[contains(@class,'el-select-dropdown') and not(contains(@style,'display: none'))]"
            "//li[contains(@class,'el-select-dropdown__item') and contains(@class,'selected')])[1]",
            "(//div[contains(@class,'el-select-dropdown') and not(contains(@style,'display: none'))]"
            "//li[.//span[contains(@class,'el-checkbox__input') and contains(@class,'is-checked')]])[1]",
            "(//div[contains(@class,'el-select-dropdown') and not(contains(@style,'display: none'))]"
            "//li[.//input[@type='checkbox' and (@checked='checked' or @checked or @aria-checked='true')]])[1]",
        ]
        for xpath in option_xpaths:
            try:
                elems = browser.find_elements(By.XPATH, xpath)
                for elem in elems:
                    if elem.is_displayed():
                        return elem
            except WebDriverException:
                continue
        return None

    def close_dropdown(self, browser):
        try:
            browser.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
        except WebDriverException:
            pass

    def confirm_dropdown_filter(self, browser):
        confirm_btn = WebDriverWait(browser, 10).until(
            EC.element_to_be_clickable((
                By.XPATH,
                "(//button[contains(@class,'el-button--primary') and (.//span[normalize-space()='确定'] or normalize-space()='确定')])[last()]"
            ))
        )
        browser.execute_script("arguments[0].click();", confirm_btn)
        time.sleep(0.8)

    def find_dropdown_search_input(self, browser):
        xpath = "//input[contains(@class,'el-select-pop__input') and @placeholder='搜索内容']"

        def visible_search_input(driver):
            elems = driver.find_elements(By.XPATH, xpath)
            for elem in reversed(elems):
                if not elem.is_displayed() or not elem.is_enabled():
                    continue
                rect = driver.execute_script(
                    "const r = arguments[0].getBoundingClientRect();"
                    "return {width:r.width,height:r.height,top:r.top,left:r.left};",
                    elem
                )
                if rect["width"] > 0 and rect["height"] > 0:
                    return elem
            return False

        return WebDriverWait(browser, 10).until(visible_search_input)

    def input_dropdown_search_text(self, browser, search_input, text):
        try:
            browser.execute_script("arguments[0].click(); arguments[0].focus();", search_input)
            search_input.send_keys(Keys.CONTROL, "a")
            search_input.send_keys(Keys.BACKSPACE)
            search_input.send_keys(text)
            return
        except WebDriverException:
            browser.execute_script(
                """
                const input = arguments[0];
                const value = arguments[1];
                input.value = value;
                input.dispatchEvent(new Event('input', {bubbles: true}));
                input.dispatchEvent(new Event('change', {bubbles: true}));
                """,
                search_input,
                text
            )
            time.sleep(0.5)

    def find_store_select_root(self, browser):
        xpath_list = [
            "//span[contains(@class,'fake-placeholder') and normalize-space()='店铺']/ancestor::div[contains(@class,'el-select')][1]",
            "//span[contains(@class,'fake-placeholder') and normalize-space()='店铺']/ancestor::div[contains(@class,'filter-item')][1]//div[contains(@class,'el-select')][1]",
            "//span[contains(@class,'el-select__tags-text') and starts-with(normalize-space(),'[eBay].')]/ancestor::div[contains(@class,'el-select')][1]",
            "//span[contains(@class,'el-tag') and contains(normalize-space(),'[eBay].')]/ancestor::div[contains(@class,'el-select')][1]",
        ]
        for xpath in xpath_list:
            try:
                elems = browser.find_elements(By.XPATH, xpath)
                for elem in elems:
                    if elem.is_displayed() and self.is_store_select(elem):
                        return elem
            except WebDriverException:
                continue
        raise TimeoutException("无法定位店铺搜索选择框")

    def is_store_select(self, elem):
        try:
            placeholders = elem.find_elements(
                By.XPATH,
                ".//span[contains(@class,'fake-placeholder') and normalize-space()='店铺']"
            )
            if len(placeholders) > 0:
                return True
            selected_store_tags = elem.find_elements(
                By.XPATH,
                ".//span[contains(@class,'el-select__tags-text') and starts-with(normalize-space(),'[eBay].')]"
                "|.//span[contains(@class,'el-tag') and contains(normalize-space(),'[eBay].')]"
            )
            return len(selected_store_tags) > 0
        except WebDriverException:
            return False

    def activate_store_search_input(self, browser, select_root):
        self.open_store_dropdown(browser, select_root)
        time.sleep(0.5)
        input_xpaths = [
            ".//input[contains(@class,'el-select__input')]",
            ".//input[contains(@class,'el-input__inner')]",
            ".//input",
        ]
        for xpath in input_xpaths:
            try:
                inputs = select_root.find_elements(By.XPATH, xpath)
                for input_elem in inputs:
                    browser.execute_script(
                        "arguments[0].removeAttribute('readonly');"
                        "arguments[0].style.display='inline-block';"
                        "arguments[0].style.visibility='visible';"
                        "arguments[0].focus();",
                        input_elem
                    )
                    if input_elem.is_enabled():
                        return input_elem
            except WebDriverException:
                continue

        active_elem = browser.switch_to.active_element
        if active_elem is not None:
            return active_elem
        raise TimeoutException("无法激活店铺搜索输入框")

    def open_store_dropdown(self, browser, select_root):
        browser.execute_script("arguments[0].scrollIntoView({block:'center'});", select_root)
        click_scripts = [
            "const root = arguments[0]; (root.querySelector('.el-input') || root).click();",
            "const root = arguments[0]; (root.querySelector('.el-input__inner') || root.querySelector('input') || root).click();",
            "const root = arguments[0]; (root.querySelector('.el-select__caret') || root).click();",
        ]
        for script in click_scripts:
            browser.execute_script(script, select_root)
            time.sleep(0.4)
            if self.has_visible_dropdown_search_input(browser):
                return

    def has_visible_dropdown_search_input(self, browser):
        try:
            elems = browser.find_elements(
                By.XPATH,
                "//input[contains(@class,'el-select-pop__input') and @placeholder='搜索内容']"
            )
            for elem in elems:
                if elem.is_displayed() and elem.is_enabled():
                    rect = browser.execute_script(
                        "const r = arguments[0].getBoundingClientRect(); return {width:r.width,height:r.height};",
                        elem
                    )
                    if rect["width"] > 0 and rect["height"] > 0:
                        return True
        except WebDriverException:
            pass
        return False

    def click_query_button_if_exists(self, browser):
        try:
            buttons = browser.find_elements(By.XPATH, "//button[.//span[normalize-space()='查询'] or normalize-space()='查询']")
            for button in buttons:
                if button.is_displayed() and button.is_enabled():
                    browser.execute_script("arguments[0].click();", button)
                    return
        except WebDriverException:
            pass

    def get_result_count(self, browser):
        total = self.get_pagination_total(browser)
        if total is not None:
            return total
        return self.get_visible_table_row_count(browser)

    def get_pagination_total(self, browser):
        try:
            elems = browser.find_elements(By.XPATH, "//*[contains(@class,'el-pagination__total')]")
            for elem in elems:
                text = (elem.text or "").strip()
                if not elem.is_displayed() or text == "":
                    continue
                match = re.search(r"\d+", text)
                if match:
                    return int(match.group())
        except WebDriverException:
            pass
        return None

    def get_visible_table_row_count(self, browser):
        count_js = """
            const rows = Array.from(document.querySelectorAll('table.vxe-table--body tr.vxe-body--row'));
            return rows.filter(row => {
                const rect = row.getBoundingClientRect();
                const style = window.getComputedStyle(row);
                return rect.width > 0 && rect.height > 0 && style.display !== 'none' && style.visibility !== 'hidden';
            }).length;
        """
        try:
            return int(browser.execute_script(count_js) or 0)
        except WebDriverException:
            return 0

    def select_all_current_page(self, browser):
        checkbox_xpaths = [
            "(//th[not(contains(@class,'fixed--hidden'))]//span[@title='全选/取消'])[1]",
            "(//th[not(contains(@class,'fixed--hidden'))]//*[contains(@class,'vxe-checkbox--icon')])[1]",
            "(//span[contains(@class,'vxe-checkbox--unchecked-icon')])[1]",
        ]
        for xpath in checkbox_xpaths:
            try:
                checkbox = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
                if checkbox.is_displayed():
                    browser.execute_script("arguments[0].click();", checkbox)
                    time.sleep(0.5)
                    return
            except (TimeoutException, StaleElementReferenceException, WebDriverException):
                continue
        raise TimeoutException("无法定位全选复选框")

    def click_delete_button(self, browser):
        delete_btn = WebDriverWait(browser, 10).until(
            EC.presence_of_element_located((
                By.XPATH,
                "(//button[not(contains(@class,'el-button--danger')) and (.//span[normalize-space()='删除'] or normalize-space()='删除')])[1]"
            ))
        )
        browser.execute_script("arguments[0].click();", delete_btn)
        time.sleep(0.5)

    def confirm_delete(self, browser):
        confirm_xpaths = [
            "//*[contains(@class,'el-message-box') or contains(@class,'el-dialog') or contains(@class,'el-popover')]"
            "//button[contains(@class,'el-button--danger') and (.//span[normalize-space()='删除'] or normalize-space()='删除')]",
            "(//button[contains(@class,'el-button--danger') and (.//span[normalize-space()='删除'] or normalize-space()='删除')])[last()]",
            "(//button[.//span[normalize-space()='确定'] or normalize-space()='确定'])[last()]",
        ]
        for xpath in confirm_xpaths:
            try:
                btn = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                browser.execute_script("arguments[0].click();", btn)
                return
            except (TimeoutException, StaleElementReferenceException, WebDriverException):
                continue
        raise TimeoutException("无法定位确认删除按钮")

    def wait_deleted(self, browser, before_count):
        end_time = time.time() + 60
        while time.time() < end_time:
            self.wait_page_idle(browser)
            current_count = self.get_result_count(browser)
            if current_count < before_count:
                return
            if current_count == 0:
                return
            if not self.is_select_all_checked(browser):
                return
            time.sleep(1)
        raise TimeoutException("等待删除刷新超时")

    def is_select_all_checked(self, browser):
        checked_xpath = "(//th[not(contains(@class,'fixed--hidden'))]//*[contains(@class,'vxe-checkbox--checked-icon')])[1]"
        try:
            checked_icons = browser.find_elements(By.XPATH, checked_xpath)
            return any(icon.is_displayed() for icon in checked_icons)
        except WebDriverException:
            return False

    def wait_page_idle(self, browser, timeout=30):
        def page_not_loading(driver):
            loading_js = """
                const nodes = Array.from(document.querySelectorAll('.el-loading-mask,.vxe-loading'));
                return nodes.every(node => {
                    const style = window.getComputedStyle(node);
                    const rect = node.getBoundingClientRect();
                    return style.display === 'none' || style.visibility === 'hidden' || rect.width === 0 || rect.height === 0;
                });
            """
            return driver.execute_script(loading_js)

        try:
            WebDriverWait(browser, timeout).until(page_not_loading)
        except TimeoutException:
            self.logger.warning("等待页面加载结束超时，继续执行后续检测")


if __name__ == "__main__":
    result = DeleteLxEbayPairedMskuJob("删除领星ebay已配对msku").run()
    print(result)
